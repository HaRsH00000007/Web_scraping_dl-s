# ============================================================
#  phase2.py — Fetch DL details from idcard.store
#  Run: python phase2.py
#  Pre-req: Chrome running via launch_chrome.bat (port 9222)
#           Logged in to idcard.store
# ============================================================

import os, re, time, logging, requests
from datetime import datetime
from openpyxl import load_workbook, Workbook
from bs4 import BeautifulSoup

from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
from webdriver_manager.chrome import ChromeDriverManager

os.environ["WDM_LOCAL"] = "1"
os.environ["WDM_LOG"]   = "0"

logging.basicConfig(level=logging.INFO, format="%(message)s")
logger = logging.getLogger("phase2")

# ── Config ───────────────────────────────────────────────────
API_URL    = "https://api.idcard.store/free/free_dl"
CDN_BASE   = "https://idmaker.mfcdn.in/"
AUTH_TOKEN = "Bearer 45b9ca25-4fe9-4aeb-bf4e-9190dcec0961"
TARGET_URL = "https://idcard.store/u/free/driving_licence"
DELAY      = 2.0   # seconds between records

HEADERS = {
    "Authorization": AUTH_TOKEN,
    "Accept":        "application/json, text/plain, */*",
    "Origin":        "https://idcard.store",
    "Referer":       "https://idcard.store/",
    "User-Agent":    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
}

COLUMNS = [
    "DL Number", "Name", "Son/Daughter/Wife of", "Date of Birth",
    "Present Address", "Mobile Number",
    "Initial Issue Date", "Initial Issuing Office",
    "Last Endorsed Date", "Last Endorsed Office",
    "Last Completed Transaction",
    "Non-Transport From", "Non-Transport To",
    "Transport From", "Transport To",
    "COV Category", "Class of Vehicle", "COV Issue Date",
    "Photo URL", "Status"
]

# ── Helpers ──────────────────────────────────────────────────

def normalise_dob(dob):
    dob = str(dob).strip()
    if re.match(r"^\d{2}-\d{2}-\d{4}$", dob): return dob
    m = re.match(r"^(\d{4})-(\d{2})-(\d{2})$", dob)
    if m: return f"{m.group(3)}-{m.group(2)}-{m.group(1)}"
    m = re.match(r"^(\d{2})/(\d{2})/(\d{4})$", dob)
    if m: return f"{m.group(1)}-{m.group(2)}-{m.group(3)}"
    return dob

def read_excel(path):
    wb = load_workbook(path)
    ws = wb.active
    headers = [str(c.value).strip() if c.value else "" for c in ws[1]]
    records = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        rec = dict(zip(headers, row))
        dl  = str(rec.get("DL Number","") or "").strip()
        dob = str(rec.get("Date of Birth","") or "").strip()
        if dl and dob and dl != "None":
            records.append({"dl": dl, "dob": normalise_dob(dob)})
    return records

def init_output(path):
    wb = Workbook()
    ws = wb.active
    ws.title = "DL Details"
    ws.append(COLUMNS)
    wb.save(path)

def append_output(path, row):
    wb = load_workbook(path)
    ws = wb.active
    ws.append([row.get(c, "") for c in COLUMNS])
    wb.save(path)

def get_done(path):
    done = set()
    if not os.path.exists(path): return done
    wb = load_workbook(path)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]: done.add(str(row[0]).strip())
    return done

def connect_driver():
    opts = Options()
    opts.add_experimental_option("debuggerAddress", "127.0.0.1:9222")
    driver = webdriver.Chrome(
        service=Service(ChromeDriverManager().install()),
        options=opts
    )
    for handle in driver.window_handles:
        driver.switch_to.window(handle)
        if "idcard.store" in driver.current_url:
            return driver
    return driver

# ── Core fetch ───────────────────────────────────────────────

def fetch_html_path(dl, dob):
    """POST to API, return CDN html path or None."""
    try:
        resp = requests.post(
            API_URL,
            data={"relation": "DL No", "dl": dl, "dob": dob},
            headers=HEADERS,
            timeout=15
        )
        if resp.status_code == 401:
            return None, "AUTH_EXPIRED"
        if resp.status_code != 200:
            return None, f"API_{resp.status_code}"
        cards = resp.json().get("cards", [])
        if not cards:
            return None, "NO_CARD"
        path = cards[0].get("html", "")
        return (path if path else None), ("OK" if path else "NO_PATH")
    except Exception as e:
        return None, f"API_ERR:{str(e)[:40]}"

def parse_html(html, dl):
    """Parse all fields from the CDN HTML page."""
    soup = BeautifulSoup(html, "html.parser")
    data = {"DL Number": dl}

    rows = soup.find_all("tr")
    for row in rows:
        cols = row.find_all("td")

        if len(cols) == 2:
            key = cols[0].get_text(strip=True).rstrip(":")
            val = cols[1].get_text(strip=True)
            if not key or not val: continue
            kl = key.lower()
            if   "name" in kl and not any(x in kl for x in ["son","daughter","wife","father"]):
                data["Name"] = val
            elif any(x in kl for x in ["son","daughter","wife","father"]):
                data["Son/Daughter/Wife of"] = val
            elif "birth" in kl:
                data["Date of Birth"] = val
            elif "address" in kl:
                data["Present Address"] = val
            elif "mobile" in kl or "phone" in kl:
                data["Mobile Number"] = val
            elif "initial issue date" in kl:
                data["Initial Issue Date"] = val
            elif "initial issuing" in kl:
                data["Initial Issuing Office"] = val
            elif "last endorsed date" in kl:
                data["Last Endorsed Date"] = val
            elif "last endorsed office" in kl:
                data["Last Endorsed Office"] = val
            elif "last completed" in kl or "transaction" in kl:
                data["Last Completed Transaction"] = val

        elif len(cols) == 3:
            v0 = cols[0].get_text(strip=True)
            v1 = cols[1].get_text(strip=True)
            v2 = cols[2].get_text(strip=True)
            kl = v0.lower()
            if v0 and v1 and "cov" not in kl and "category" not in kl and "class" not in kl:
                if not data.get("COV Category"):
                    data["COV Category"]     = v0
                    data["Class of Vehicle"] = v1
                    data["COV Issue Date"]   = v2

        elif len(cols) >= 4:
            k1 = cols[0].get_text(strip=True).lower()
            v1 = cols[1].get_text(strip=True).replace("From:","").strip()
            v2 = cols[3].get_text(strip=True).replace("To:","").strip()
            if "non-transport" in k1:
                data["Non-Transport From"] = v1
                data["Non-Transport To"]   = v2
            elif "transport" in k1 and "non" not in k1:
                data["Transport From"] = v1
                data["Transport To"]   = v2

    # Photo
    img = soup.find("img")
    if img and img.get("src"):
        data["Photo URL"] = img["src"][:300]

    return data

def process_record(driver, dl, dob):
    """Full pipeline: API → CDN fetch → parse."""
    result = {"DL Number": dl, "Status": ""}

    # Step 1: Get HTML path from API
    html_path, status = fetch_html_path(dl, dob)
    if not html_path:
        result["Status"] = status
        if status == "AUTH_EXPIRED":
            raise Exception("AUTH_EXPIRED")
        return result

    # Step 2: Fetch HTML via Selenium (CDN, no Cloudflare)
    html_url = CDN_BASE + html_path
    logger.info(f"  → {html_url}")

    html = ""
    for attempt in range(4):          # up to 4 attempts
        try:
            driver.get(html_url)
            # Wait until a table appears (max 8s)
            WebDriverWait(driver, 8).until(
                EC.presence_of_element_located((By.TAG_NAME, "table"))
            )
            html = driver.page_source
            if "<table" in html:
                break
        except TimeoutException:
            logger.info(f"  Attempt {attempt+1}: table not ready, retrying...")
            time.sleep(2)
        except Exception as e:
            logger.warning(f"  Attempt {attempt+1} error: {e}")
            time.sleep(2)

    if not html or "<table" not in html:
        result["Status"] = "NO_TABLE"
        driver.get(TARGET_URL)
        time.sleep(2)
        return result

    # Step 3: Parse
    data = parse_html(html, dl)

    if data.get("Name"):
        result.update(data)
        result["Status"] = "OK"
    else:
        result["Status"] = "NO_DATA"

    # Navigate back to form
    driver.get(TARGET_URL)
    time.sleep(2)
    return result

# ── Main ─────────────────────────────────────────────────────

def main():
    print("\n" + "="*60)
    print("  Phase 2 — DL Details Fetcher")
    print("="*60)
    print("""
  Before running:
  1. Double-click launch_chrome.bat
  2. Log in to idcard.store in that Chrome
  """)

    input_file = input("  Enter Phase 1 Excel path: ").strip().strip('"')
    if not os.path.exists(input_file):
        print(f"  File not found: {input_file}")
        return

    records = read_excel(input_file)
    print(f"  Loaded {len(records)} records.")

    os.makedirs("output", exist_ok=True)
    ts          = datetime.now().strftime("%Y-%m-%d_%H%M%S")
    output_path = os.path.join("output", f"DL_Details_{ts}.xlsx")
    init_output(output_path)
    done = get_done(output_path)

    print("  Connecting to Chrome...")
    try:
        driver = connect_driver()
        print(f"  Connected: {driver.current_url}")
    except Exception as e:
        print(f"  Cannot connect: {e}")
        return

    print(f"  Output: {output_path}")
    print("  Starting... (Ctrl+C to stop)\n")

    saved = failed = 0

    for i, rec in enumerate(records):
        dl  = rec["dl"]
        dob = rec["dob"]

        if dl in done:
            print(f"  [{i+1}/{len(records)}] SKIP: {dl}")
            continue

        try:
            result = process_record(driver, dl, dob)
        except Exception as e:
            if "AUTH_EXPIRED" in str(e):
                print("\n  !! AUTH TOKEN EXPIRED — update AUTH_TOKEN in phase2.py !!")
                break
            result = {"DL Number": dl, "Status": f"ERROR:{str(e)[:50]}"}

        append_output(output_path, result)

        if result["Status"] == "OK":
            saved += 1
            print(f"  [{i+1}/{len(records)}] OK  : {dl} | {result.get('Name','?')} | {result.get('Date of Birth','?')}")
        else:
            failed += 1
            print(f"  [{i+1}/{len(records)}] FAIL: {dl} | {result['Status']}")

        time.sleep(DELAY)

    print(f"\n  Done. {saved} saved, {failed} failed → {output_path}")


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\n  Stopped.")