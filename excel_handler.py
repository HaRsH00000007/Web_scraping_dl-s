# ============================================================
#  excel_handler.py  —  Write results to Excel
# ============================================================

import os
from openpyxl import Workbook, load_workbook
from datetime import datetime
from typing import Dict
from config import OUTPUT_DIR

COLUMNS = ["DL Number", "Name", "Date of Birth", "Application No", "State", "RTO"]


def get_output_path(state_name: str) -> str:
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    timestamp  = datetime.now().strftime("%Y-%m-%d_%H%M%S")
    safe_state = state_name.replace(" ", "_")
    return os.path.join(OUTPUT_DIR, f"{safe_state}_{timestamp}.xlsx")


def init_output_file(path: str, logger) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "DL Data"
    ws.append(COLUMNS)
    wb.save(path)
    logger.info(f"Output file created: {path}")


def append_row(path: str, data: Dict, logger) -> None:
    """Append one result row — crash-safe, writes immediately."""
    try:
        wb = load_workbook(path)
        ws = wb.active
        ws.append([data.get(col, "") for col in COLUMNS])
        wb.save(path)
        logger.debug(f"Saved: {data.get('DL Number')} | {data.get('Name')}")
    except Exception as e:
        logger.error(f"Excel write error: {e}")